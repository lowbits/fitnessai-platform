#!/usr/bin/env python3
"""
Batch remove backgrounds from exercise illustrations and save as transparent WebP.
Processes illustrations that were already uploaded to R2, creating transparent versions.

Usage:
    python3 scripts/remove_bg_batch.py --input-dir /tmp/exercise-illustrations --output-dir /tmp/exercise-illustrations-transparent
    python3 scripts/remove_bg_batch.py --report storage/app/exercise-smart-match.xlsx --illustration-dir "/Volumes/Seagate/ILLUSTRATIONS" --output-dir /tmp/transparent --mobile-only
"""

import argparse
import os
import sys
import time
from pathlib import Path

from PIL import Image
from rembg import remove, new_session


def normalize(name):
    """Normalize exercise name to match the key format used in the upload command."""
    import re
    name = name.lower().strip()
    name = re.sub(r'_?(female|male)$', '', name, flags=re.IGNORECASE)
    name = name.replace('-', ' ').replace('_', ' ').replace('\u2013', ' ').replace('\u2014', ' ')
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def find_illustration(illustration_dir, bundle_key):
    """Find the illustration file for a bundle key (female preferred, male fallback)."""
    for category_dir in Path(illustration_dir).iterdir():
        if not category_dir.is_dir():
            continue
        for ext in ['jpeg', 'jpg', 'png']:
            # Look for files, skip frame 2 (ending in digit before extension)
            for f in category_dir.glob(f'*.{ext}'):
                basename = f.stem
                # Skip frame 2
                if basename[-1].isdigit() and not basename.endswith('female') and not basename.endswith('male'):
                    continue

                file_key = normalize(basename)
                if file_key == bundle_key:
                    # Check if female
                    is_female = basename.lower().endswith('_female') or basename.lower().endswith('female')
                    return str(f), is_female
    return None, False


def process_image(input_path, output_path, session, quality=85):
    """Remove background and save as transparent WebP."""
    img = Image.open(input_path)
    output = remove(img, session=session)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    output.save(output_path, 'WEBP', quality=quality)

    orig_size = os.path.getsize(input_path)
    new_size = os.path.getsize(output_path)
    return orig_size, new_size


def load_report(report_path, mobile_only=False):
    """Load the smart-match report and return exercises to process."""
    from openpyxl import load_workbook
    wb = load_workbook(report_path, read_only=True)
    ws = wb.active

    exercises = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        action, mobile_usage, total_usage, db_name, bundle_name, db_id = row[:6]
        match_type = row[6] if len(row) > 6 else ''

        if action != 'UPDATE_MEDIA':
            continue
        if mobile_only and (not mobile_usage or int(mobile_usage) == 0):
            continue
        if not bundle_name:
            continue

        # Get slug from db_name (simple slugify)
        slug = db_name.lower().strip()
        slug = slug.replace(' ', '-').replace('(', '').replace(')', '').replace("'", '')
        slug = slug.replace('/', '-').replace('--', '-').strip('-')

        exercises.append({
            'db_id': db_id,
            'db_name': db_name,
            'bundle_name': bundle_name,
            'bundle_key': normalize(bundle_name),
            'slug': slug,
            'mobile_usage': int(mobile_usage) if mobile_usage else 0,
        })

    wb.close()
    return exercises


def main():
    parser = argparse.ArgumentParser(description='Batch remove backgrounds from exercise illustrations')
    parser.add_argument('--report', help='Path to smart-match XLSX report')
    parser.add_argument('--illustration-dir', required=True, help='Path to illustrations directory')
    parser.add_argument('--output-dir', required=True, help='Output directory for transparent images')
    parser.add_argument('--mobile-only', action='store_true', help='Only process mobile-used exercises')
    parser.add_argument('--quality', type=int, default=85, help='WebP quality (default: 85)')
    parser.add_argument('--limit', type=int, help='Limit number of images to process')
    parser.add_argument('--skip-existing', action='store_true', help='Skip if output file already exists')

    args = parser.parse_args()

    # Load exercise list from report
    if args.report:
        exercises = load_report(args.report, args.mobile_only)
    else:
        print("--report is required")
        sys.exit(1)

    if args.limit:
        exercises = exercises[:args.limit]

    print(f"Exercises to process: {len(exercises)}")

    # Create rembg session (reuses model across images)
    print("Loading AI model...")
    session = new_session("u2net")

    # Index illustrations
    print("Indexing illustrations...")
    illustration_index = {}
    ill_dir = Path(args.illustration_dir)
    for category_dir in ill_dir.iterdir():
        if not category_dir.is_dir():
            continue
        for f in category_dir.glob('*.*'):
            if f.suffix.lower() not in ['.jpeg', '.jpg', '.png']:
                continue
            basename = f.stem
            # Skip frame 2
            if len(basename) > 0 and basename[-1].isdigit():
                # Check it's actually a frame number not part of the name
                test_name = basename[:-1]
                if test_name and (test_name.endswith('_female') or test_name.endswith('_male') or test_name[-1] != '_'):
                    continue

            is_female = '_female' in basename.lower() or basename.lower().endswith('female')
            key = normalize(basename)

            if key not in illustration_index:
                illustration_index[key] = {'male': None, 'female': None}

            if is_female:
                illustration_index[key]['female'] = str(f)
            else:
                illustration_index[key]['male'] = str(f)

    print(f"Illustrations indexed: {len(illustration_index)}")

    processed = 0
    skipped = 0
    total_orig = 0
    total_new = 0
    start_time = time.time()

    for i, ex in enumerate(exercises):
        bundle_key = ex['bundle_key']
        slug = ex['slug']
        output_path = os.path.join(args.output_dir, f"{slug}.webp")

        if args.skip_existing and os.path.exists(output_path):
            skipped += 1
            continue

        # Find illustration (female preferred)
        ill_info = illustration_index.get(bundle_key, {})
        input_path = ill_info.get('female') or ill_info.get('male')

        if not input_path:
            skipped += 1
            continue

        try:
            orig_size, new_size = process_image(input_path, output_path, session, args.quality)
            total_orig += orig_size
            total_new += new_size
            processed += 1

            elapsed = time.time() - start_time
            rate = processed / elapsed if elapsed > 0 else 0
            eta = (len(exercises) - i - 1) / rate if rate > 0 else 0

            print(f"  [{i+1}/{len(exercises)}] {ex['db_name']} -> {slug}.webp "
                  f"({orig_size/1024:.0f}KB -> {new_size/1024:.0f}KB) "
                  f"[{rate:.1f}/s, ETA: {eta/60:.0f}m]")
        except Exception as e:
            print(f"  ERROR [{i+1}/{len(exercises)}] {ex['db_name']}: {e}")
            skipped += 1

    elapsed = time.time() - start_time
    print(f"\n=== DONE ===")
    print(f"Processed: {processed}")
    print(f"Skipped: {skipped}")
    print(f"Time: {elapsed/60:.1f} minutes")
    if total_orig > 0:
        print(f"Size: {total_orig/1024/1024:.1f}MB -> {total_new/1024/1024:.1f}MB ({(1-total_new/total_orig)*100:.0f}% reduction)")


if __name__ == '__main__':
    main()
